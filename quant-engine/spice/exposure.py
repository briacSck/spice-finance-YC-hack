"""Exposure extractor: ordinary monthly accounting (CSV/XLSX) -> commodity exposure
map + hedge priorities.

Implements EXPOSURE_AGENT_PLAN.md. Reads the tidy CSV emitted by cashplan
(scripts/export_cost_csv.py) or a generated cash-plan workbook, reuses
spice.quant for cost-share, margin stress and VaR, and adds mapping confidence,
monthly volatility, a hedge-priority score, evidence rows, and JSON/Markdown
reports.

    CSV in (cashplan)  ->  map -> metrics -> stress -> score  ->  JSON + Markdown
                              (reuses spice.commodities, spice.quant)

CLI:  python -m spice.exposure <cashplan.xlsx|costs.csv> [--json out.json] [--md out.md] [--llm]
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import statistics
import sys
from dataclasses import dataclass, field
from pathlib import Path

from . import commodities, quant
from .schema import Company, CostLine

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# Scoring calibration: score = 50% cost share + 50% normalized historical cost volatility.
COST_SHARE_FULL = 0.30      # 30% of cost base -> full cost-share points
VOL_FULL = 0.50             # coefficient of variation of historical monthly cost -> full vol points
SHOCK = 0.20                # the headline stress shock (informational only, not part of score)


# --- ingestion ---------------------------------------------------------------
@dataclass
class Row:
    business: str
    archetype: str
    month: str       # YYYY-MM
    kind: str        # expense | revenue
    category: str
    line: str
    underlying_raw: str
    amount: float
    source_row: int


def load_rows(csv_path: str | Path) -> list[Row]:
    rows: list[Row] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        for i, r in enumerate(csv.DictReader(f), start=2):  # row 1 = header
            try:
                amount = float(r.get("amount") or 0.0)
            except ValueError:
                amount = 0.0
            rows.append(
                Row(
                    business=(r.get("business") or "").strip(),
                    archetype=(r.get("archetype") or "").strip(),
                    month=(r.get("period_month") or "").strip(),
                    kind=(r.get("kind") or "expense").strip().lower(),
                    category=(r.get("category") or "").strip(),
                    line=(r.get("line") or "").strip(),
                    underlying_raw=(r.get("underlying") or "").strip(),
                    amount=amount,
                    source_row=i,
                )
            )
    return rows


# --- mapping -----------------------------------------------------------------
@dataclass
class Mapping:
    ticker: str | None
    confidence: float
    rule: str


def map_row(r: Row) -> Mapping:
    """Trust the underlying tag first; fall back to keyword inference."""
    t = commodities.ticker_for_underlying(r.underlying_raw)
    if t:
        return Mapping(t, 0.95, f"tag '{r.underlying_raw}' -> {t}")
    t = commodities.infer_ticker(f"{r.line} {r.category}")
    if t:
        return Mapping(t, 0.75, f"keyword -> {t}")
    return Mapping(None, 0.0, "unmapped")


# --- analysis ----------------------------------------------------------------
@dataclass
class Exposure:
    ticker: str
    underlying: str
    total_spend: float
    expense_share: float
    revenue_share: float
    mapping_confidence: float
    monthly_volatility: float       # coefficient of variation of monthly spend
    price_volatility: float         # annualised commodity price vol (INSEE if available, else registry)
    price_vol_source: str           # "INSEE BDM <code>" or registry fallback reason
    shock20_margin_points_lost: float
    hedge_notional: float
    score: int
    rank: int
    recommendation: str
    evidence: list[dict] = field(default_factory=list)


@dataclass
class Report:
    business: dict
    financial_summary: dict
    exposures: list[Exposure]
    unmapped: dict
    var: dict
    llm: dict | None = None


def _num(value, default: float = 0.0) -> float:
    try:
        if value is None or isinstance(value, str):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _start_year(path: Path) -> int:
    match = re.search(r"(20\d{2})", path.stem)
    return int(match.group(1)) if match else 2026


def _clean_cost_label(value: str) -> str:
    text = (value or "").strip()
    if text.lower().startswith("cost (") and text.endswith(")"):
        return text[6:-1]
    return text.replace(" - market price", "").strip()


def _load_cashplan_workbook(path: str | Path) -> list[Row]:
    """Read the generated cash-plan .xlsx directly.

    Generated workbooks contain formulas, but openpyxl does not evaluate them.
    This reconstructs revenue and expense rows from the hypothesis cells,
    market-price rows, and monthly headers instead of relying on cached formula
    values.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx input. Install requirements-cashplan.txt.") from exc

    p = Path(path)
    wb = load_workbook(p, data_only=False)
    required = {"Hypotheses", "Revenus", "Depenses detaillees"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise SystemExit(f"Unsupported workbook: missing sheets {sorted(missing)}")

    hyp = wb["Hypotheses"]
    rev = wb["Revenus"]
    exp = wb["Depenses detaillees"]

    title = str(hyp["A1"].value or p.stem)
    business = title.split(" - ")[0].strip() or p.stem
    archetype = ""
    subtitle = str(hyp["A2"].value or "")
    match = re.search(r"Archetype:\s*([^|]+)", subtitle)
    if match:
        archetype = match.group(1).strip()

    annual_revenue = _num(hyp["C4"].value)
    revenue_growth = _num(hyp["C13"].value)
    purchase_inflation = _num(hyp["C14"].value)
    seasonality = [_num(hyp.cell(17, c).value, 1.0) for c in range(3, 15)]
    season_sum = sum(seasonality) or 1.0
    start_year = _start_year(p)

    month_cols: list[tuple[int, str, int, int]] = []
    for offset, col in enumerate(range(8, exp.max_column + 1)):
        if not exp.cell(4, col).value:
            continue
        year_index = offset // 12
        month_index = offset % 12
        month = f"{start_year + year_index:04d}-{month_index + 1:02d}"
        month_cols.append((col, month, year_index, month_index))

    rows: list[Row] = []
    source_row = 2

    # Revenue streams.
    for rr in range(5, rev.max_row + 1):
        label = str(rev.cell(rr, 1).value or "").strip()
        if not label or label.startswith("TOTAL") or label.startswith("Control") or label.startswith("TVA") or label.startswith("CA TTC"):
            continue
        share = _num(rev.cell(rr, 2).value)
        for _, month, year_index, month_index in month_cols:
            amount = annual_revenue * share * (seasonality[month_index] / season_sum) * ((1 + revenue_growth) ** year_index)
            rows.append(Row(business, archetype, month, "revenue", "Revenue", label, "", round(amount, 2), source_row))
            source_row += 1

    # Expense lines. Rows named "Cost (...)" use the previous market-price row.
    for er in range(5, exp.max_row + 1):
        category = str(exp.cell(er, 1).value or "").strip()
        label = str(exp.cell(er, 2).value or "").strip()
        if not category or category.startswith("SOUS-TOTAL") or category.startswith("TOTAL") or not label:
            continue
        if label.endswith(" - market price"):
            continue

        quantity = _num(exp.cell(er, 3).value, 1.0)
        unit_price = _num(exp.cell(er, 5).value)
        underlying = str(exp.cell(er, 6).value or "").strip()
        clean_label = _clean_cost_label(label)
        is_cost_from_market = label.lower().startswith("cost (")
        market_row = er - 1 if is_cost_from_market else None

        for col, month, year_index, month_index in month_cols:
            formula = str(exp.cell(er, col).value or "")
            if is_cost_from_market and market_row:
                price = _num(exp.cell(market_row, col).value, unit_price)
                season = seasonality[month_index] if "$17" in formula else 1.0
                amount = quantity * season * price
            else:
                season = seasonality[month_index] if "$17" in formula else 1.0
                amount = quantity * season * unit_price * ((1 + purchase_inflation) ** year_index)
            rows.append(Row(business, archetype, month, "expense", category, clean_label, underlying, round(amount, 2), er))

    return rows


def load_input(path: str | Path) -> list[Row]:
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_cashplan_workbook(p)
    return load_rows(p)


_insee_provider = None


def _get_insee_provider():
    global _insee_provider
    if _insee_provider is None:
        from cashplan.data_sources import CommodityDataProvider

        _insee_provider = CommodityDataProvider(cache_dir=_REPO_ROOT / ".cache" / "cashplan", online=True)
    return _insee_provider


INSEE_HISTORY_YEARS = 10
INSEE_HISTORY_MONTHS = INSEE_HISTORY_YEARS * 12


def _annualized_return_vol(values: list[float]) -> float | None:
    if len(values) < 3:
        return None
    returns = [values[i] / values[i - 1] - 1 for i in range(1, len(values)) if values[i - 1]]
    if len(returns) < 2:
        return None
    return statistics.pstdev(returns) * (12 ** 0.5)


def _insee_price_vol(ticker: str) -> tuple[float | None, str]:
    """Real annualised price volatility from the last 10 years of INSEE BDM
    monthly prices only - no other data source (Alpha Vantage/World Bank/synthetic)."""
    code = commodities.insee_code(ticker)
    if not code:
        return None, "no INSEE series mapped for this ticker"
    try:
        series = _get_insee_provider()._fetch_insee_series(code)
    except Exception as exc:
        return None, f"INSEE fetch failed: {exc}"
    if not series or not series.points:
        return None, f"INSEE series {code} unavailable (no network/cache)"
    values = [p.value for p in series.points if math.isfinite(p.value)][-INSEE_HISTORY_MONTHS:]
    vol = _annualized_return_vol(values)
    if vol is None:
        return None, f"not enough INSEE {code} data points to compute volatility"
    years_used = len(values) / 12
    return vol, f"INSEE BDM {code} ({years_used:.0f}y monthly)"


def analyze(rows: list[Row]) -> Report:
    expenses = [r for r in rows if r.kind == "expense"]
    revenues = [r for r in rows if r.kind == "revenue"]
    months = sorted({r.month for r in rows if r.month})
    n_months = max(1, len(months))
    years = n_months / 12.0

    total_expense = sum(r.amount for r in expenses)
    total_revenue = sum(r.amount for r in revenues)
    margin = total_revenue - total_expense
    margin_pct = margin / total_revenue if total_revenue else 0.0

    # aggregate expenses by ticker
    by_ticker_spend: dict[str, float] = {}
    by_ticker_monthly: dict[str, dict[str, float]] = {}
    by_ticker_conf: dict[str, float] = {}
    by_ticker_rows: dict[str, list[Row]] = {}
    unmapped_total = 0.0
    unmapped_rows: list[Row] = []

    for r in expenses:
        m = map_row(r)
        if not m.ticker:
            unmapped_total += r.amount
            unmapped_rows.append(r)
            continue
        by_ticker_spend[m.ticker] = by_ticker_spend.get(m.ticker, 0.0) + r.amount
        monthly = by_ticker_monthly.setdefault(m.ticker, {})
        monthly[r.month] = monthly.get(r.month, 0.0) + r.amount
        by_ticker_conf[m.ticker] = max(by_ticker_conf.get(m.ticker, 0.0), m.confidence)
        by_ticker_rows.setdefault(m.ticker, []).append(r)

    # build a Company (annualised) so we can reuse quant.var + recommend_hedges
    cost_lines = [
        CostLine(label=commodities.en_name(t), annual_amount=spend / years, commodity=t)
        for t, spend in by_ticker_spend.items()
    ]
    if unmapped_total:
        cost_lines.append(CostLine("Other (labour, rent, …)", unmapped_total / years, None))
    company = Company(
        name=rows[0].business if rows else "Business",
        revenue=total_revenue / years if total_revenue else 0.0,
        cash=0.0,
        cost_lines=cost_lines,
    )
    hedges = {h.commodity: h.notional for h in quant.recommend_hedges(company)}

    exposures: list[Exposure] = []
    for t, spend in sorted(by_ticker_spend.items(), key=lambda kv: kv[1], reverse=True):
        series = [by_ticker_monthly[t].get(m, 0.0) for m in months]
        mean = statistics.fmean(series) if series else 0.0
        monthly_vol = (statistics.pstdev(series) / mean) if mean else 0.0
        price_vol, price_vol_source = _insee_price_vol(t)
        if price_vol is None:
            price_vol_source = f"registry fallback ({price_vol_source})"
            price_vol = commodities.vol(t)
        # +20% shock on this commodity only -> margin points lost
        extra_cost = spend * SHOCK
        points_lost = (extra_cost / total_revenue * 100) if total_revenue else 0.0

        # score = 50% cost share + 50% normalized market price volatility (not
        # spend volatility, which conflates price moves with purchased quantity).
        score = round(
            50 * min((spend / total_expense) / COST_SHARE_FULL, 1.0)
            + 50 * min(price_vol / VOL_FULL, 1.0)
        )
        evidence = [
            {
                "source_row": r.source_row,
                "description": r.line,
                "amount": round(r.amount, 2),
                "matched_rule": map_row(r).rule,
            }
            for r in sorted(by_ticker_rows[t], key=lambda r: r.amount, reverse=True)[:3]
        ]
        exposures.append(
            Exposure(
                ticker=t,
                underlying=commodities.en_name(t),
                total_spend=round(spend, 2),
                expense_share=round(spend / total_expense, 4) if total_expense else 0.0,
                revenue_share=round(spend / total_revenue, 4) if total_revenue else 0.0,
                mapping_confidence=round(by_ticker_conf.get(t, 0.0), 2),
                monthly_volatility=round(monthly_vol, 4),
                price_volatility=round(price_vol, 4),
                price_vol_source=price_vol_source,
                shock20_margin_points_lost=round(points_lost, 2),
                hedge_notional=round(hedges.get(t, 0.0), 2),
                score=score,
                rank=0,
                recommendation="Pending LLM decision",
                evidence=evidence,
            )
        )

    # Rank by score (highest first) - this ranking is what the LLM decides on.
    exposures.sort(key=lambda e: e.score, reverse=True)
    for i, e in enumerate(exposures, start=1):
        e.rank = i

    var_before = quant.var(company)
    var_after = quant.var(company, residual_basis=0.15)

    return Report(
        business={
            "name": company.name,
            "period_start": months[0] if months else None,
            "period_end": months[-1] if months else None,
            "months_analyzed": len(months),
        },
        financial_summary={
            "average_monthly_revenue": round(total_revenue / n_months, 2),
            "average_monthly_expenses": round(total_expense / n_months, 2),
            "average_margin_pct": round(margin_pct, 4),
            "mapped_expense_share": round(
                (total_expense - unmapped_total) / total_expense, 4
            ) if total_expense else 0.0,
        },
        exposures=exposures,
        unmapped={
            "expense_share": round(unmapped_total / total_expense, 4) if total_expense else 0.0,
            "top_rows": [
                {"source_row": r.source_row, "description": r.line, "amount": round(r.amount, 2)}
                for r in sorted(unmapped_rows, key=lambda r: r.amount, reverse=True)[:5]
            ],
        },
        var={
            "annual_margin_var_pct_before": round(var_before.pct_revenue, 4),
            "annual_margin_var_pct_after_hedge": round(var_after.pct_revenue, 4),
            "eur_before": round(var_before.eur, 2),
            "eur_after": round(var_after.eur, 2),
        },
    )


def _llm_payload(report: Report) -> dict:
    top = [
        {
            "underlying": e.underlying,
            "ticker": e.ticker,
            "rank": e.rank,
            "score": e.score,
            "expense_share": e.expense_share,
            "revenue_share": e.revenue_share,
            "price_volatility": e.price_volatility,
            "shock20_margin_points_lost": e.shock20_margin_points_lost,
            "evidence": e.evidence,
        }
        for e in report.exposures[:6]
    ]
    return {
        "business": report.business,
        "financial_summary": report.financial_summary,
        "top_exposures": top,
        "unmapped": report.unmapped,
        "var": report.var,
    }


def _llm_prompt(report: Report) -> str:
    return (
        "You are the narrative layer AND the hedging decision-maker for Spice's "
        "exposure agent. The numbers are already computed and must not be changed "
        "or recomputed - quote them exactly as given. Return strict JSON with keys: "
        "executive_angle, owner_message, demo_script, ambiguity_notes, "
        "hedge_decisions.\n\n"
        "hedge_decisions must be a JSON array, one entry per item in top_exposures, "
        "each entry an object with exactly: ticker (copy from input), decision (one "
        "of 'Hedge recommended' or 'Do not hedge yet'), reasoning (one short "
        "sentence grounded in that item's own expense_share, price_volatility and "
        "shock20_margin_points_lost - not a generic statement). Use the rank and "
        "score only as a starting point, not a hard rule: a high-rank item with a "
        "low cost share might still not be worth hedging, and you may deviate from "
        "the ranking order if the underlying numbers justify it - explain why in "
        "the reasoning when you do.\n\n"
        "Write for a bakery/restaurant/shop owner with no finance background. Ban "
        "generic phrases like 'indicate a need for hedging strategies to mitigate "
        "potential margin impacts' - that explains nothing. Instead, for each top "
        "exposure, spell out the causal chain with the actual numbers: which raw "
        "ingredient or input it is, what share of costs it represents, and what "
        "happens to their pocket if that commodity's price jumps 20% (use "
        "shock20_margin_points_lost converted to euros via average_monthly_revenue, "
        "e.g. '2.8 margin points on €26,485/month revenue is roughly -€740/month "
        "gone, on top of an already thin margin'). Mention the unmapped expense "
        "share and why it limits the analysis (e.g. salaries/rent aren't "
        "commodity-linked so they're out of scope, not ignored by mistake). Define "
        "'hedge' once in plain words if you use the word (e.g. 'locking in a price "
        "now instead of paying whatever the market charges later'). No jargon left "
        "unexplained, no claim without a number behind it. Do not give financial "
        "advice. Data:\n"
        + json.dumps(_llm_payload(report), ensure_ascii=False)
    )


def _parse_llm_json(text: str) -> dict:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
    return {"raw": text}


def _safe_error(exc: Exception) -> str:
    text = str(exc)
    text = re.sub(r"sk-[A-Za-z0-9_-]+", "sk-***", text)
    text = re.sub(r"sk-proj-[A-Za-z0-9_-]+", "sk-proj-***", text)
    text = re.sub(r"[A-Za-z0-9_-]{6}\*{3,}[A-Za-z0-9_-]{4,}", "***redacted***", text)
    return text


def load_local_env() -> None:
    """Load local ignored env files without overriding existing variables."""
    roots = [Path.cwd(), Path.cwd().parent, Path(__file__).resolve().parents[2]]
    for root in roots:
        for name in [".env.local", ".env"]:
            path = root / name
            if not path.exists():
                continue
            override = name == ".env.local"
            for raw_line in path.read_text(encoding="utf-8").splitlines():
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and (override or key not in os.environ):
                    os.environ[key] = value


def _resolve_llm_provider(provider: str) -> str | None:
    if provider != "auto":
        return provider
    if os.getenv("OPENAI_API_KEY"):
        return "openai"
    if os.getenv("ANTHROPIC_API_KEY"):
        return "anthropic"
    return None


def _apply_hedge_decisions(report: Report) -> None:
    """Let the LLM's per-ticker hedge_decisions override the placeholder recommendation."""
    if not report.llm or report.llm.get("status") != "ok":
        return
    decisions = report.llm.get("hedge_decisions")
    if not isinstance(decisions, list):
        return
    by_ticker = {d.get("ticker"): d for d in decisions if isinstance(d, dict)}
    for e in report.exposures:
        d = by_ticker.get(e.ticker)
        if d and d.get("decision"):
            e.recommendation = d["decision"]


def smooth_with_llm(report: Report, provider: str, model: str | None) -> Report:
    """Add an optional narrative layer without changing deterministic metrics."""
    resolved = _resolve_llm_provider(provider)
    if not resolved:
        report.llm = {
            "status": "skipped",
            "reason": "OPENAI_API_KEY or ANTHROPIC_API_KEY is not set; deterministic report is still complete.",
        }
        return report

    prompt = _llm_prompt(report)
    system = (
        "You polish a commodity exposure report for an SMB owner. "
        "Never invent numbers, venues, trades, or hidden supply-chain inference."
    )

    if resolved == "openai":
        if not os.getenv("OPENAI_API_KEY"):
            report.llm = {
                "status": "skipped",
                "provider": "openai",
                "reason": "OPENAI_API_KEY is not set.",
            }
            return report
        try:
            from openai import OpenAI
        except ImportError:
            report.llm = {
                "status": "skipped",
                "provider": "openai",
                "reason": "openai package is not installed; install quant-engine requirements to enable --llm.",
            }
            return report
        model = model or os.getenv("OPENAI_MODEL", "gpt-4o-mini")
        try:
            client = OpenAI()
            response = client.chat.completions.create(
                model=model,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.2,
            )
            text = response.choices[0].message.content or "{}"
            report.llm = {"status": "ok", "provider": "openai", "model": model, **_parse_llm_json(text)}
            _apply_hedge_decisions(report)
        except Exception as exc:
            report.llm = {"status": "error", "provider": "openai", "model": model, "reason": _safe_error(exc)}
        return report

    if resolved != "anthropic":
        report.llm = {"status": "skipped", "reason": f"Unknown LLM provider: {resolved}"}
        return report

    if not os.getenv("ANTHROPIC_API_KEY"):
        report.llm = {
            "status": "skipped",
            "provider": "anthropic",
            "reason": "ANTHROPIC_API_KEY is not set.",
        }
        return report
    try:
        from anthropic import Anthropic
    except ImportError:
        report.llm = {
            "status": "skipped",
            "provider": "anthropic",
            "reason": "anthropic package is not installed; install quant-engine requirements to enable --llm.",
        }
        return report

    model = model or os.getenv("ANTHROPIC_MODEL", "claude-opus-4-8")
    try:
        client = Anthropic()
        msg = client.messages.create(
            model=model,
            max_tokens=900,
            system=system,
            messages=[{"role": "user", "content": prompt}],
        )
        text = "\n".join(
            block.text for block in msg.content if getattr(block, "type", None) == "text"
        ).strip()
        report.llm = {"status": "ok", "provider": "anthropic", "model": model, **_parse_llm_json(text)}
        _apply_hedge_decisions(report)
    except Exception as exc:
        report.llm = {"status": "error", "provider": "anthropic", "model": model, "reason": _safe_error(exc)}
    return report


# --- outputs -----------------------------------------------------------------
def to_json(report: Report) -> str:
    payload = {
        "business": report.business,
        "financial_summary": report.financial_summary,
        "exposures": [e.__dict__ for e in report.exposures],
        "unmapped": report.unmapped,
        "var": report.var,
    }
    if report.llm:
        payload["llm"] = report.llm
    return json.dumps(payload, ensure_ascii=False, indent=2)


def to_markdown(report: Report) -> str:
    b, fs = report.business, report.financial_summary
    lines = [
        f"# Commodity exposure report — {b['name']}",
        "",
        f"Period {b['period_start']} to {b['period_end']} ({b['months_analyzed']} months). "
        f"Avg monthly revenue €{fs['average_monthly_revenue']:,.0f}, "
        f"avg margin {fs['average_margin_pct'] * 100:.1f}%, "
        f"{fs['mapped_expense_share'] * 100:.0f}% of expenses mapped to a commodity.",
        "",
    ]
    lines += [
        "## Hedge priorities (ranked by score)",
        "",
        "| Rank | Underlying | Cost share | Price vol | Vol source | +20% shock (margin pts) | Score | Recommendation |",
        "|--:|---|--:|--:|---|--:|--:|---|",
    ]
    for e in report.exposures:
        lines.append(
            f"| {e.rank} | {e.underlying} ({e.ticker}) | {e.expense_share * 100:.0f}% | "
            f"{e.price_volatility * 100:.0f}% | {e.price_vol_source} | {e.shock20_margin_points_lost:.1f} | "
            f"{e.score} | {e.recommendation} |"
        )
    lines += [
        "",
        f"Annual margin VaR: {report.var['annual_margin_var_pct_before'] * 100:.1f}% of revenue "
        f"unhedged → {report.var['annual_margin_var_pct_after_hedge'] * 100:.1f}% hedged.",
        "",
        "<details><summary>How the score and recommendation are computed</summary>",
        "",
        "Score (0-100) is a weighted sum, each term capped at its max:",
        f"- 50 pts: cost share, capped at {COST_SHARE_FULL * 100:.0f}% of total expenses",
        f"- 50 pts: annualised price volatility of the underlying commodity, capped at "
        f"{VOL_FULL * 100:.0f}%",
        "",
        "Price volatility is computed from real INSEE BDM monthly price series "
        "(std of month-over-month returns, annualised) when a series is mapped for "
        "that ticker - no other external source (Alpha Vantage, World Bank, synthetic) "
        "is used. If INSEE has no series for a ticker or the fetch/cache both fail, it "
        "falls back to a static registry estimate, flagged in the 'Vol source' column.",
        "",
        "This score sets the ranking only. The final Recommendation per line is decided "
        "by the LLM narrative layer (see Hedge decisions below), which can deviate from "
        "the ranking when the underlying numbers justify it. Without an LLM call "
        "(--llm not set or no API key), the Recommendation stays 'Pending LLM decision'.",
        "",
        "</details>",
        "",
    ]
    decisions = (report.llm or {}).get("hedge_decisions") if report.llm and report.llm.get("status") == "ok" else None
    if isinstance(decisions, list) and decisions:
        lines += ["## Hedge decisions (AI reasoning)", ""]
        for d in decisions:
            if isinstance(d, dict) and d.get("ticker"):
                lines.append(f"- **{d['ticker']}** — {d.get('decision', '')}: {d.get('reasoning', '')}")
        lines.append("")
    lines += [
        "## Evidence (top contributing lines)",
        "",
    ]
    for e in report.exposures:
        lines.append(f"**{e.underlying}** — confidence {e.mapping_confidence:.2f}")
        for ev in e.evidence:
            lines.append(f"- row {ev['source_row']}: {ev['description']} (€{ev['amount']:,.0f}) — {ev['matched_rule']}")
        lines.append("")
    if report.unmapped["top_rows"]:
        lines.append(f"## Unmapped ({report.unmapped['expense_share'] * 100:.0f}% of expenses)")
        for ev in report.unmapped["top_rows"]:
            lines.append(f"- row {ev['source_row']}: {ev['description']} (€{ev['amount']:,.0f})")
        lines.append("")
    if report.llm:
        lines += ["## Findings explained (AI)", ""]
        if report.llm.get("status") == "ok":
            for key in ["executive_angle", "owner_message", "demo_script", "ambiguity_notes"]:
                if report.llm.get(key):
                    lines.append(f"**{key.replace('_', ' ').title()}**: {report.llm[key]}")
                    lines.append("")
        else:
            lines.append(f"{report.llm.get('status')}: {report.llm.get('reason')}")
            lines.append("")
    lines.append("> Risk analysis, not financial advice. v1 maps direct inputs only, no multi-hop supply chain.")
    return "\n".join(lines)


# --- CLI ---------------------------------------------------------------------
def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # Windows console is cp1252
    except Exception:
        pass
    ap = argparse.ArgumentParser(description="Spice exposure extractor")
    ap.add_argument("input", help="generated cash-plan .xlsx or tidy cost CSV from cashplan")
    ap.add_argument("--json", dest="json_out", default=None)
    ap.add_argument("--md", dest="md_out", default=None)
    ap.add_argument("--llm", action="store_true", help="add an optional LLM narrative layer")
    ap.add_argument("--llm-provider", choices=["auto", "openai", "anthropic"], default="auto")
    ap.add_argument("--llm-model", default=None)
    args = ap.parse_args()

    load_local_env()
    report = analyze(load_input(args.input))
    if args.llm:
        report = smooth_with_llm(report, args.llm_provider, args.llm_model)
    js = to_json(report)
    md = to_markdown(report)

    if args.json_out:
        Path(args.json_out).write_text(js, encoding="utf-8")
    if args.md_out:
        Path(args.md_out).write_text(md, encoding="utf-8")
    if not args.json_out and not args.md_out:
        print(md)
        print("\n--- JSON ---\n")
        print(js)


if __name__ == "__main__":
    main()
