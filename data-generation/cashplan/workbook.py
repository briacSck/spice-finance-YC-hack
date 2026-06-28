from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from cashplan.business_profiles import BusinessProfile, CostLine
from cashplan.data_sources import CommodityDataProvider, PriceSeries


HYP = "Hypotheses"
REV = "Revenus"
EXP = "Depenses detaillees"
PLAN = "Plan de tresorerie"
MARKET = "Market data"
SOURCES = "Benchmarks & sources"

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

TITLE_FILL = "1F2937"
SECTION_FILL = "E5E7EB"
HEADER_FILL = "374151"
INPUT_FILL = "DDEBFF"
FORMULA_FILL = "F3F4F6"
LINK_FILL = "E0F2FE"
SOURCE_FILL = "FEF3C7"
TOTAL_FILL = "DCFCE7"
WARN_FILL = "FEE2E2"
WHITE = "FFFFFF"
TEXT = "111827"
MUTED = "6B7280"


@dataclass(frozen=True)
class MonthSpec:
    index: int
    year_index: int
    month_index: int
    iso_month: str
    label: str
    workbook_col: int
    seasonality_col: int


@dataclass(frozen=True)
class MarketSeriesEntry:
    line_name: str
    underlying: str | None
    market_code: str
    unit: str
    unit_price: float
    series: PriceSeries


def build_cash_plan_workbook(
    profile: BusinessProfile,
    output_path: Path,
    data_provider: CommodityDataProvider,
    start_year: int = 2026,
    years: int = 3,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    months = _build_months(start_year, years)
    wb = Workbook()
    wb.remove(wb.active)

    ws_hyp = wb.create_sheet(HYP)
    ws_rev = wb.create_sheet(REV)
    ws_exp = wb.create_sheet(EXP)
    ws_plan = wb.create_sheet(PLAN)
    ws_market = wb.create_sheet(MARKET)
    ws_sources = wb.create_sheet(SOURCES)

    _build_hypotheses(ws_hyp, profile)
    revenue_rows = _build_revenues(ws_rev, profile, months)
    category_subtotals, used_series = _build_expenses(ws_exp, profile, months, data_provider)
    _build_cash_plan(ws_plan, profile, months, revenue_rows, category_subtotals)
    _build_market_data(ws_market, used_series)
    _build_sources(ws_sources, profile, data_provider)

    for ws in wb.worksheets:
        _apply_sheet_defaults(ws)

    wb.save(output_path)
    return output_path


def _build_months(start_year: int, years: int) -> list[MonthSpec]:
    months: list[MonthSpec] = []
    for index in range(years * 12):
        year_index = index // 12
        month_index = index % 12
        year = start_year + year_index
        workbook_col = 3 + index
        seasonality_col = 3 + month_index
        months.append(
            MonthSpec(
                index=index,
                year_index=year_index,
                month_index=month_index,
                iso_month=f"{year:04d}-{month_index + 1:02d}",
                label=f"{MONTH_NAMES[month_index]} A{year_index + 1}",
                workbook_col=workbook_col,
                seasonality_col=seasonality_col,
            )
        )
    return months


def _build_hypotheses(ws: Worksheet, profile: BusinessProfile) -> None:
    ws["A1"] = f"{profile.business_type.title()} - Hypotheses"
    ws["A2"] = (
        f"Metier: {profile.description} | Archetype: {profile.archetype} | "
        f"CA cible A1: {profile.annual_revenue:,.0f} EUR | Blue cells are editable"
    )

    assumptions = [
        ("Chiffre d'affaires annuel HT - Annee 1 (EUR)", profile.annual_revenue),
        ("Taux de TVA moyen sur les ventes", profile.vat_sales),
        ("Taux de TVA moyen sur les achats", profile.vat_purchases),
        ("Tresorerie disponible au 1er janvier A1 (EUR)", profile.opening_cash),
        ("Cotisations sociales dirigeant - par trimestre A1 (EUR)", profile.owner_social_quarterly),
        ("Remboursement emprunt materiel - mensualite fixe (EUR)", profile.loan_monthly),
        ("Part du CA encaissee le mois meme", 1.0),
        ("Part du CA encaissee a M-1", 0.0),
        ("Part du CA encaissee a M-2", 0.0),
        ("Taux de croissance du CA - par an", profile.annual_revenue_growth),
        ("Taux d'inflation des prix d'achat - par an", profile.purchase_inflation),
    ]
    for row, (label, value) in enumerate(assumptions, start=4):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value=value)
        _style_input(ws.cell(row=row, column=3))

    ws["D12"] = "Control should equal 100%"
    ws["E12"] = "=SUM(C10:C12)"
    _style_formula(ws["E12"])

    ws["A16"] = "Month pattern repeated every year"
    ws["A17"] = "Seasonality index (1 = average month)"
    ws["A18"] = "Month weight in year (normalized)"
    for idx, month_name in enumerate(MONTH_NAMES, start=3):
        ws.cell(row=16, column=idx, value=month_name)
        ws.cell(row=17, column=idx, value=profile.seasonality[idx - 3])
        ws.cell(row=18, column=idx, value=f"={get_column_letter(idx)}17/SUM($C$17:$N$17)")
        _style_input(ws.cell(row=17, column=idx))
        _style_formula(ws.cell(row=18, column=idx))

    notes = [
        "v1 note: ratios and quantities are synthetic benchmark assumptions unless a source/API is explicitly listed.",
        "Commodity-linked rows preserve realistic monthly variation. External series are normalized to the model unit price.",
        "INSEE support is intentionally configurable by URL/series because exact indicators depend on the sector and use case.",
    ]
    for row, note in enumerate(notes, start=21):
        ws.cell(row=row, column=1, value=note)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    legend = [
        ("Blue", "Editable assumption"),
        ("Gray", "Formula calculated automatically"),
        ("Light blue", "Link to another sheet"),
        ("Yellow", "Source/API note"),
    ]
    ws["A25"] = "Legend"
    for row, (name, meaning) in enumerate(legend, start=26):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=meaning)


def _build_revenues(ws: Worksheet, profile: BusinessProfile, months: list[MonthSpec]) -> dict[str, int]:
    ws["A1"] = f"{profile.business_type.title()} - Revenus mensuels"
    ws["A2"] = "Revenue streams x annual mix x seasonality x annual growth."
    ws["A4"] = "Source de revenus"
    ws["B4"] = "% du CA annuel"
    for month in months:
        ws.cell(row=4, column=month.workbook_col, value=month.label)
        if month.month_index == 0:
            ws.cell(row=3, column=month.workbook_col, value=f"ANNEE {month.year_index + 1}")

    first_stream_row = 5
    for row_offset, stream in enumerate(profile.revenue_streams):
        row = first_stream_row + row_offset
        ws.cell(row=row, column=1, value=stream.name)
        ws.cell(row=row, column=2, value=stream.share)
        _style_input(ws.cell(row=row, column=2))
        for month in months:
            col_letter = get_column_letter(month.workbook_col)
            season_col = get_column_letter(month.seasonality_col)
            formula = (
                f"='{HYP}'!$C$4*$B{row}*'{HYP}'!{season_col}$18*"
                f"(1+'{HYP}'!$C$13)^{month.year_index}"
            )
            ws.cell(row=row, column=month.workbook_col, value=formula)
            _style_formula(ws.cell(row=row, column=month.workbook_col))

    total_row = first_stream_row + len(profile.revenue_streams)
    control_row = total_row + 1
    vat_row = total_row + 2
    ttc_row = total_row + 3

    ws.cell(row=total_row, column=1, value="TOTAL CA HT")
    ws.cell(row=control_row, column=1, value="Control sum revenue mix")
    ws.cell(row=control_row, column=2, value=f"=SUM(B{first_stream_row}:B{total_row - 1})")
    ws.cell(row=vat_row, column=1, value="TVA collectee sur les ventes")
    ws.cell(row=ttc_row, column=1, value="CA TTC encaissable")
    for month in months:
        col = get_column_letter(month.workbook_col)
        ws.cell(row=total_row, column=month.workbook_col, value=f"=SUM({col}{first_stream_row}:{col}{total_row - 1})")
        ws.cell(row=vat_row, column=month.workbook_col, value=f"={col}{total_row}*'{HYP}'!$C$5")
        ws.cell(row=ttc_row, column=month.workbook_col, value=f"={col}{total_row}+{col}{vat_row}")
        _style_formula(ws.cell(row=total_row, column=month.workbook_col))
        _style_formula(ws.cell(row=vat_row, column=month.workbook_col))
        _style_formula(ws.cell(row=ttc_row, column=month.workbook_col))

    for row in [4, total_row, control_row, vat_row, ttc_row]:
        _style_section_row(ws, row, 2 + len(months))
    _style_formula(ws.cell(row=control_row, column=2))
    return {"total_ht": total_row, "vat": vat_row, "ttc": ttc_row}


def _build_expenses(
    ws: Worksheet,
    profile: BusinessProfile,
    months: list[MonthSpec],
    data_provider: CommodityDataProvider,
) -> tuple[dict[str, int], list[MarketSeriesEntry]]:
    ws["A1"] = f"{profile.business_type.title()} - Depenses detaillees"
    ws["A2"] = "Quantity x seasonality x unit price x inflation = monthly cost. Market rows use monthly price series."

    headers = ["Categorie", "Poste", "Qte base/mois A1", "Unite", "Prix unitaire A1", "Underlying", "Source prix"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=header)
    for month in months:
        ws.cell(row=4, column=7 + month.index + 1, value=month.label)
        if month.month_index == 0:
            ws.cell(row=3, column=7 + month.index + 1, value=f"ANNEE {month.year_index + 1}")

    row = 5
    current_category: str | None = None
    category_cost_rows: dict[str, list[int]] = {}
    category_subtotals: dict[str, int] = {}
    used_series: list[MarketSeriesEntry] = []

    def close_category(category: str, next_row: int) -> int:
        subtotal_row = next_row
        ws.cell(row=subtotal_row, column=1, value=f"SOUS-TOTAL {category}")
        for month in months:
            col = get_column_letter(8 + month.index)
            cost_refs = [f"{col}{cost_row}" for cost_row in category_cost_rows.get(category, [])]
            ws.cell(row=subtotal_row, column=8 + month.index, value=f"=SUM({','.join(cost_refs)})" if cost_refs else 0)
            _style_formula(ws.cell(row=subtotal_row, column=8 + month.index))
        _style_total_row(ws, subtotal_row, 7 + len(months))
        category_subtotals[category] = subtotal_row
        return subtotal_row + 1

    for line in profile.cost_lines:
        if current_category and line.category != current_category:
            row = close_category(current_category, row)
        if line.category != current_category:
            current_category = line.category
            category_cost_rows.setdefault(current_category, [])

        if line.market_code:
            price_row = row
            cost_row = row + 1
            months_iso = [month.iso_month for month in months]
            series = data_provider.get_series(line.market_code, months_iso, line.unit_price)
            used_series.append(
                MarketSeriesEntry(
                    line_name=line.name,
                    underlying=line.underlying,
                    market_code=line.market_code,
                    unit=line.unit,
                    unit_price=line.unit_price,
                    series=series,
                )
            )

            ws.cell(row=price_row, column=1, value=line.category)
            ws.cell(row=price_row, column=2, value=f"{line.name} - market price")
            ws.cell(row=price_row, column=3, value=line.monthly_quantity)
            ws.cell(row=price_row, column=4, value=line.unit)
            ws.cell(row=price_row, column=5, value=f"=AVERAGE(H{price_row}:{get_column_letter(7 + len(months))}{price_row})")
            ws.cell(row=price_row, column=6, value=line.underlying)
            ws.cell(row=price_row, column=7, value=f"{series.source}; {series.price_basis}")
            _style_source(ws.cell(row=price_row, column=7))
            _style_input(ws.cell(row=price_row, column=3))
            for month in months:
                value = series.points[month.index].value
                ws.cell(row=price_row, column=8 + month.index, value=value)
                _style_source(ws.cell(row=price_row, column=8 + month.index))

            ws.cell(row=cost_row, column=1, value=line.category)
            ws.cell(row=cost_row, column=2, value=f"Cost ({line.name})")
            ws.cell(row=cost_row, column=3, value=line.monthly_quantity)
            ws.cell(row=cost_row, column=4, value=line.unit)
            ws.cell(row=cost_row, column=5, value=f"=E{price_row}")
            ws.cell(row=cost_row, column=6, value=line.underlying)
            ws.cell(row=cost_row, column=7, value="Cost = quantity x seasonality x monthly market price")
            _style_input(ws.cell(row=cost_row, column=3))
            _style_formula(ws.cell(row=cost_row, column=5))
            for month in months:
                col = get_column_letter(8 + month.index)
                if line.seasonality_linked:
                    season_col = get_column_letter(month.seasonality_col)
                    season_multiplier = f"*'{HYP}'!{season_col}$17"
                else:
                    season_multiplier = ""
                formula = f"=$C{cost_row}{season_multiplier}*{col}{price_row}"
                ws.cell(row=cost_row, column=8 + month.index, value=formula)
                _style_formula(ws.cell(row=cost_row, column=8 + month.index))
            category_cost_rows[line.category].append(cost_row)
            row += 2
        else:
            ws.cell(row=row, column=1, value=line.category)
            ws.cell(row=row, column=2, value=line.name)
            ws.cell(row=row, column=3, value=line.monthly_quantity)
            ws.cell(row=row, column=4, value=line.unit)
            ws.cell(row=row, column=5, value=line.unit_price)
            ws.cell(row=row, column=6, value=line.underlying)
            ws.cell(row=row, column=7, value=line.source)
            _style_input(ws.cell(row=row, column=3))
            _style_input(ws.cell(row=row, column=5))
            _style_source(ws.cell(row=row, column=7))
            for month in months:
                season_multiplier = f"*'{HYP}'!{get_column_letter(month.seasonality_col)}$17" if line.seasonality_linked else ""
                formula = (
                    f"=$C{row}*{season_multiplier[1:] if season_multiplier else '1'}*"
                    f"$E{row}*(1+'{HYP}'!$C$14)^{month.year_index}"
                )
                ws.cell(row=row, column=8 + month.index, value=formula)
                _style_formula(ws.cell(row=row, column=8 + month.index))
            category_cost_rows[line.category].append(row)
            row += 1

    if current_category:
        row = close_category(current_category, row)

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAL GENERAL DEPENSES")
    for month in months:
        col = get_column_letter(8 + month.index)
        subtotal_refs = [f"{col}{subtotal_row}" for subtotal_row in category_subtotals.values()]
        ws.cell(row=total_row, column=8 + month.index, value=f"=SUM({','.join(subtotal_refs)})")
        _style_formula(ws.cell(row=total_row, column=8 + month.index))
    _style_total_row(ws, total_row, 7 + len(months))

    control_row = total_row + 1
    ws.cell(row=control_row, column=1, value="Total depenses A1 / CA A1")
    first_month_col = get_column_letter(8)
    last_a1_col = get_column_letter(8 + 11)
    ws.cell(row=control_row, column=3, value=f"=SUM({first_month_col}{total_row}:{last_a1_col}{total_row})/'{HYP}'!$C$4")
    _style_formula(ws.cell(row=control_row, column=3))

    _style_section_row(ws, 4, 7 + len(months))
    return category_subtotals, used_series


def _build_cash_plan(
    ws: Worksheet,
    profile: BusinessProfile,
    months: list[MonthSpec],
    revenue_rows: dict[str, int],
    category_subtotals: dict[str, int],
) -> None:
    ws["A1"] = f"{profile.business_type.title()} - Plan de tresorerie mensuel"
    ws["A2"] = "Opening cash + collections - disbursements = ending cash."
    ws["A4"] = "Mois"
    for month in months:
        ws.cell(row=4, column=month.workbook_col, value=month.label)
        if month.month_index == 0:
            ws.cell(row=3, column=month.workbook_col, value=f"ANNEE {month.year_index + 1}")

    rows = {
        "opening": 5,
        "collections_section": 6,
        "client_collections": 7,
        "total_collections": 8,
        "disbursements_section": 9,
        "commodity_purchases": 10,
        "personnel": 11,
        "fixed": 12,
        "owner_social": 13,
        "vat_paid": 14,
        "loan": 15,
        "total_disbursements": 16,
        "variation": 17,
        "ending": 18,
        "memo_vat_collected": 20,
        "memo_vat_deductible": 21,
        "memo_vat_net": 22,
        "min_cash": 24,
    }
    labels = {
        "opening": "Tresorerie en debut de mois",
        "collections_section": "ENCAISSEMENTS",
        "client_collections": "Encaissements clients TTC",
        "total_collections": "TOTAL ENCAISSEMENTS",
        "disbursements_section": "DECAISSEMENTS",
        "commodity_purchases": "Achats variables et commodity-linked",
        "personnel": "Personnel",
        "fixed": "Charges fixes et maintenance",
        "owner_social": "Cotisations sociales dirigeant",
        "vat_paid": "TVA nette versee",
        "loan": "Remboursement emprunt materiel",
        "total_disbursements": "TOTAL DECAISSEMENTS",
        "variation": "Variation de tresorerie du mois",
        "ending": "TRESORERIE EN FIN DE MOIS",
        "memo_vat_collected": "Memo - TVA collectee du mois",
        "memo_vat_deductible": "Memo - TVA deductible du mois",
        "memo_vat_net": "Memo - TVA nette du mois",
        "min_cash": "Minimum cash over period",
    }
    for key, row in rows.items():
        ws.cell(row=row, column=1, value=labels[key])

    commodity_categories = [
        category for category in category_subtotals if category not in {"Personnel", "Fixed costs"}
    ]
    personnel_categories = [category for category in category_subtotals if category == "Personnel"]
    fixed_categories = [category for category in category_subtotals if category == "Fixed costs"]

    for month in months:
        col_num = month.workbook_col
        col = get_column_letter(col_num)
        exp_col = get_column_letter(8 + month.index)

        if month.index == 0:
            ws.cell(row=rows["opening"], column=col_num, value=f"='{HYP}'!$C$7")
        else:
            prev_col = get_column_letter(col_num - 1)
            ws.cell(row=rows["opening"], column=col_num, value=f"={prev_col}{rows['ending']}")

        ttc_row = revenue_rows["ttc"]
        current = f"'{REV}'!{col}{ttc_row}*'{HYP}'!$C$10"
        previous = f"'{REV}'!{get_column_letter(col_num - 1)}{ttc_row}*'{HYP}'!$C$11" if month.index >= 1 else "0"
        two_back = f"'{REV}'!{get_column_letter(col_num - 2)}{ttc_row}*'{HYP}'!$C$12" if month.index >= 2 else "0"
        ws.cell(row=rows["client_collections"], column=col_num, value=f"={current}+{previous}+{two_back}")
        ws.cell(row=rows["total_collections"], column=col_num, value=f"={col}{rows['client_collections']}")

        ws.cell(row=rows["commodity_purchases"], column=col_num, value=_sum_category_refs(EXP, exp_col, category_subtotals, commodity_categories))
        ws.cell(row=rows["personnel"], column=col_num, value=_sum_category_refs(EXP, exp_col, category_subtotals, personnel_categories))
        ws.cell(row=rows["fixed"], column=col_num, value=_sum_category_refs(EXP, exp_col, category_subtotals, fixed_categories))
        if month.month_index in {2, 5, 8, 11}:
            ws.cell(row=rows["owner_social"], column=col_num, value=f"='{HYP}'!$C$8*(1+'{HYP}'!$C$14)^{month.year_index}")
        else:
            ws.cell(row=rows["owner_social"], column=col_num, value=0)

        if month.month_index in {2, 5, 8, 11}:
            start_col = get_column_letter(max(3, col_num - 2))
            ws.cell(row=rows["vat_paid"], column=col_num, value=f"=MAX(0,SUM({start_col}{rows['memo_vat_net']}:{col}{rows['memo_vat_net']}))")
        else:
            ws.cell(row=rows["vat_paid"], column=col_num, value=0)

        ws.cell(row=rows["loan"], column=col_num, value=f"='{HYP}'!$C$9")
        ws.cell(row=rows["total_disbursements"], column=col_num, value=f"=SUM({col}{rows['commodity_purchases']}:{col}{rows['loan']})")
        ws.cell(row=rows["variation"], column=col_num, value=f"={col}{rows['total_collections']}-{col}{rows['total_disbursements']}")
        ws.cell(row=rows["ending"], column=col_num, value=f"={col}{rows['opening']}+{col}{rows['variation']}")
        ws.cell(row=rows["memo_vat_collected"], column=col_num, value=f"='{REV}'!{col}{revenue_rows['vat']}")
        ws.cell(row=rows["memo_vat_deductible"], column=col_num, value=f"=({col}{rows['commodity_purchases']}+{col}{rows['fixed']})*'{HYP}'!$C$6")
        ws.cell(row=rows["memo_vat_net"], column=col_num, value=f"={col}{rows['memo_vat_collected']}-{col}{rows['memo_vat_deductible']}")
        for row in rows.values():
            if row not in {rows["collections_section"], rows["disbursements_section"]}:
                _style_formula(ws.cell(row=row, column=col_num))

    first_col = get_column_letter(3)
    last_col = get_column_letter(2 + len(months))
    ws.cell(row=rows["min_cash"], column=3, value=f"=MIN({first_col}{rows['ending']}:{last_col}{rows['ending']})")
    _style_formula(ws.cell(row=rows["min_cash"], column=3))

    for row in [4, rows["collections_section"], rows["disbursements_section"], rows["total_collections"], rows["total_disbursements"], rows["ending"], rows["min_cash"]]:
        _style_section_row(ws, row, 2 + len(months))


def _build_market_data(ws: Worksheet, used_series: list[MarketSeriesEntry]) -> None:
    ws["A1"] = "Market data used by model"
    ws["A2"] = "External series are normalized to each model unit price to preserve realistic movement without unit mismatch."
    headers = ["Cost line", "Underlying", "Market code", "Source", "Model unit", "Model unit price", "Month", "Value"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=header)
    row = 5
    for entry in sorted(used_series, key=lambda item: (item.market_code, item.line_name)):
        for point in entry.series.points:
            ws.cell(row=row, column=1, value=entry.line_name)
            ws.cell(row=row, column=2, value=entry.underlying)
            ws.cell(row=row, column=3, value=entry.market_code)
            ws.cell(row=row, column=4, value=point.source)
            ws.cell(row=row, column=5, value=entry.unit)
            ws.cell(row=row, column=6, value=entry.unit_price)
            ws.cell(row=row, column=7, value=point.month)
            ws.cell(row=row, column=8, value=point.value)
            row += 1
    _style_section_row(ws, 4, 8)


def _build_sources(ws: Worksheet, profile: BusinessProfile, data_provider: CommodityDataProvider) -> None:
    ws["A1"] = f"{profile.business_type.title()} - Benchmarks & sources"
    ws["A2"] = "Ratios and API sources used to construct the synthetic cash-plan model."
    headers = ["Item", "Value retained", "Source", "Comment"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=header)

    rows = [
        ("Annual revenue A1", profile.annual_revenue, "Generated profile", "Editable in Hypotheses"),
        ("Archetype", profile.archetype, "Local inference", "Selected from business description keywords"),
        ("Commodity API mode", "online" if data_provider.online else "offline", "Generator flag", "Offline mode uses deterministic synthetic series"),
    ]
    rows.extend(profile.benchmark_notes)
    for row_idx, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3 and isinstance(value, str) and value.startswith("http"):
                ws.cell(row=row_idx, column=col_idx).hyperlink = value
                ws.cell(row=row_idx, column=col_idx).style = "Hyperlink"
    _style_section_row(ws, 4, 4)


def _sum_category_refs(sheet_name: str, col: str, category_subtotals: dict[str, int], categories: list[str]) -> str:
    refs = [f"'{sheet_name}'!{col}{category_subtotals[category]}" for category in categories if category in category_subtotals]
    if not refs:
        return 0
    return f"=SUM({','.join(refs)})"


def _apply_sheet_defaults(ws: Worksheet) -> None:
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 34,
        "B": 42,
        "C": 16,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(8, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 12

    ws.row_dimensions[1].height = 24
    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Aptos", color=TEXT)
            cell.alignment = cell.alignment.copy(vertical="center")
            if cell.row == 1:
                cell.font = Font(name="Aptos Display", bold=True, size=16, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
            elif cell.row == 2:
                cell.font = Font(name="Aptos", italic=True, size=10, color=MUTED)

    for row in [1, 2]:
        if ws.max_column > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(ws.max_column, 10))

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20


def _style_input(cell: Cell) -> None:
    cell.fill = PatternFill("solid", fgColor=INPUT_FILL)


def _style_formula(cell: Cell) -> None:
    cell.fill = PatternFill("solid", fgColor=FORMULA_FILL)


def _style_source(cell: Cell) -> None:
    cell.fill = PatternFill("solid", fgColor=SOURCE_FILL)


def _style_section_row(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL if row == 4 else SECTION_FILL)
        cell.font = Font(name="Aptos", bold=True, color=WHITE if row == 4 else TEXT)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="CBD5E1"))


def _style_total_row(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=TOTAL_FILL)
        cell.font = Font(name="Aptos", bold=True, color=TEXT)
        cell.border = Border(top=Side(style="thin", color="94A3B8"))


def add_source_comment(cell: Cell, text: str) -> None:
    cell.comment = Comment(text, "Spice")
